import openpyxl
import pandas as pd
from datetime import datetime, timedelta

# ============= 1. 读取第一个文件的数据 =============
df1 = pd.read_excel('危局强袭战怪物数据_69001-69041.xlsx')

season_data = {}
for _, row in df1.iterrows():
    season = int(row['赛季ID'])
    if season not in season_data:
        season_data[season] = []
    hp = row['生命值_Lv29']
    if isinstance(hp, str):
        hp = int(hp.replace(',', ''))
    season_data[season].append({
        'name_full': row['怪物名称'],
        'name_short': row['阶段标识'],
        'hp': hp
    })

# ============= 2. 完整名称到简写的映射 =============
name_short_map = {
    '牲鬼·布林格': '牲鬼',
    '「霸主侵蚀体·庞培」': '白庞培',
    '恶名·冥宁芙': '双子',
    '自律强袭单位·「提丰·破坏者型」': '提丰',
    '死路屠夫': '屠夫',
    '恶名·死路屠夫': '恶名屠夫',
    '未知复合侵蚀体': '侵蚀体',
    '恶名·庞培': '红庞培',
    '秽息司祭': '司祭',
    '秽息妖鬼·名可名': '名可名',
    '「亵渎者」': '亵渎者',
    '彷徨猎手': '彷徨',
    '魇缚者·叶释渊': '叶释渊',
    '太初梦魇·「始主」': '莎拉',
    '叛律孤歌·薇斯珀': '叛律孤歌',
    '猎血清道夫': '清道夫',
    '异变能量体': '异变能量体',
    '基塔布鲁': '基塔布鲁',
}

# 简写到完整名称的映射（用于匹配已有数据）
name_full_map = {v: k for k, v in name_short_map.items()}
# 处理别名
name_full_map['判律孤歌'] = '叛律孤歌·薇斯珀'
name_full_map['冥宁芙'] = '恶名·冥宁芙'

# ============= 3. 用openpyxl打开第二个文件 =============
wb = openpyxl.load_workbook('危局2.7图鉴版.xlsx')
ws = wb['总表']

# ============= 4. 更新已有赛季的血量 =============
print("更新已有数据...")
update_count = 0

for season_id in range(69001, 69036):
    # 计算起始行号
    start_row = 3 + (season_id - 69001) * 4
    
    # 该赛季的3个怪物行
    monster_rows = [start_row, start_row + 1, start_row + 2]
    
    if season_id not in season_data:
        print(f"  警告: 赛季 {season_id} 在源文件中不存在")
        continue
    
    monsters = season_data[season_id]
    if len(monsters) != 3:
        print(f"  警告: 赛季 {season_id} 有 {len(monsters)} 个怪物，预期3个")
        continue
    
    for i, (r, monster) in enumerate(zip(monster_rows, monsters)):
        current_hp = ws.cell(row=r, column=3).value
        expected_hp = monster['hp']
        
        # 转换当前值为整数
        if isinstance(current_hp, str):
            try:
                current_hp = int(current_hp.replace(',', ''))
            except:
                current_hp = None
        
        if current_hp != expected_hp:
            old_val = ws.cell(row=r, column=3).value
            ws.cell(row=r, column=3).value = expected_hp
            print(f"  更新 Row {r} (赛季{season_id} 阶段{i+1}): {old_val} -> {expected_hp}")
            update_count += 1

print(f"共更新 {update_count} 个单元格")

# ============= 5. 准备69036-69041的数据 =============
print("\n准备新数据...")

# 日期推算：69035是20260424，之后每14天一个赛季
base_date = datetime(2026, 4, 24)
new_seasons = []
for season_id in range(69036, 69042):
    offset = (season_id - 69035) * 14
    date_val = base_date + timedelta(days=offset)
    date_str = date_val.strftime('%Y%m%d')
    new_seasons.append({
        'season_id': season_id,
        'date': int(date_str),
        'monsters': season_data[season_id]
    })
    print(f"  赛季 {season_id}: 日期 {date_str}")

# ============= 6. 插入69036的数据 =============
# 69036的起始行是Row 143
start_row_36 = 3 + (69036 - 69001) * 4  # = 143

# 当前Row 143: 日期
# 当前Row 144: 版本
# 需要在Row 144之后插入2行（怪物3 + 总）

print(f"\n插入69036数据到 Row {start_row_36}...")

# 填入Row 143（怪物1）
ws.cell(row=start_row_36, column=2).value = name_short_map[new_seasons[0]['monsters'][0]['name_full']]
ws.cell(row=start_row_36, column=3).value = new_seasons[0]['monsters'][0]['hp']
print(f"  Row {start_row_36}: 怪物1 = {name_short_map[new_seasons[0]['monsters'][0]['name_full']]}")

# 填入Row 144（怪物2）
ws.cell(row=start_row_36 + 1, column=2).value = name_short_map[new_seasons[0]['monsters'][1]['name_full']]
ws.cell(row=start_row_36 + 1, column=3).value = new_seasons[0]['monsters'][1]['hp']
print(f"  Row {start_row_36 + 1}: 怪物2 = {name_short_map[new_seasons[0]['monsters'][1]['name_full']]}")

# 在Row 144之后插入2行
ws.insert_rows(start_row_36 + 2, 2)

# 新Row 145: 怪物3
ws.cell(row=start_row_36 + 2, column=2).value = name_short_map[new_seasons[0]['monsters'][2]['name_full']]
ws.cell(row=start_row_36 + 2, column=3).value = new_seasons[0]['monsters'][2]['hp']
print(f"  Row {start_row_36 + 2}: 怪物3 = {name_short_map[new_seasons[0]['monsters'][2]['name_full']]}")

# 新Row 146: 总
ws.cell(row=start_row_36 + 3, column=1).value = '总'
ws.cell(row=start_row_36 + 3, column=6).value = f'=C{start_row_36}+C{start_row_36+1}+C{start_row_36+2}'
print(f"  Row {start_row_36 + 3}: 总 = =C{start_row_36}+C{start_row_36+1}+C{start_row_36+2}")

# ============= 7. 追加69037-69041的数据 =============
# 插入行后，69037的起始行变了
# 原来69037的起始行应该是 3 + (69037-69001)*4 = 147
# 但由于我们插入了2行，实际起始行是 147 + 2 = 149

current_row = start_row_36 + 4  # = 147

for season_info in new_seasons[1:]:  # 69037-69041
    season_id = season_info['season_id']
    monsters = season_info['monsters']
    date_val = season_info['date']
    
    print(f"\n追加赛季 {season_id} 到 Row {current_row}...")
    
    # 怪物1行：日期 + 怪物1
    ws.cell(row=current_row, column=1).value = date_val
    ws.cell(row=current_row, column=2).value = name_short_map[monsters[0]['name_full']]
    ws.cell(row=current_row, column=3).value = monsters[0]['hp']
    print(f"  Row {current_row}: 日期={date_val}, 怪物1={name_short_map[monsters[0]['name_full']]}")
    
    # 怪物2行
    ws.cell(row=current_row + 1, column=2).value = name_short_map[monsters[1]['name_full']]
    ws.cell(row=current_row + 1, column=3).value = monsters[1]['hp']
    print(f"  Row {current_row + 1}: 怪物2={name_short_map[monsters[1]['name_full']]}")
    
    # 怪物3行
    ws.cell(row=current_row + 2, column=2).value = name_short_map[monsters[2]['name_full']]
    ws.cell(row=current_row + 2, column=3).value = monsters[2]['hp']
    print(f"  Row {current_row + 2}: 怪物3={name_short_map[monsters[2]['name_full']]}")
    
    # 总行
    ws.cell(row=current_row + 3, column=1).value = '总'
    ws.cell(row=current_row + 3, column=6).value = f'=C{current_row}+C{current_row+1}+C{current_row+2}'
    print(f"  Row {current_row + 3}: 总")
    
    current_row += 4

# ============= 8. 保存文件 =============
output_path = '危局2.7图鉴版_已补充.xlsx'
wb.save(output_path)
print(f"\n文件已保存到: {output_path}")

