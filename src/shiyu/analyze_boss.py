import openpyxl
import numpy as np
from scipy import stats
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

# Set Chinese font
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# ============================================================
# 1. Load data
# ============================================================
wb = openpyxl.load_workbook('式舆怪物血量数据.xlsx')
ws = wb[wb.sheetnames[0]]

bosses = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    node_id = row[0]
    monster = row[3]
    hp_str = row[4]
    if hp_str is None:
        continue
    hp = int(str(hp_str).replace(',', ''))
    if hp > 12_600_000:
        bosses.append((node_id, monster, hp))

node_ids = sorted(set(b[0] for b in bosses))
monster_names = sorted(set(b[1] for b in bosses))

print(f"Nodes: {len(node_ids)}, Boss types: {len(monster_names)}")
print(f"Node range: {node_ids[0]} ~ {node_ids[-1]}")

# Build pivot: dict[monster][node_id] = hp
pivot = defaultdict(dict)
for nid, monster, hp in bosses:
    pivot[monster][nid] = hp

# Count occurrences per monster
occurrences = {m: len(pivot[m]) for m in monster_names}
print("\nOccurrences per monster:")
for m, c in sorted(occurrences.items(), key=lambda x: -x[1]):
    nodes_present = sorted(pivot[m].keys())
    print(f"  {m}: {c} times -> nodes {nodes_present}")

# ============================================================
# 2. Create Pivot Table (Excel)
# ============================================================
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

out_wb = openpyxl.Workbook()

# --- Sheet 1: Pivot Table ---
ws_pivot = out_wb.active
ws_pivot.title = "Boss HP Pivot Table"

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Header row
ws_pivot.cell(row=1, column=1, value="Monster Name")
for j, nid in enumerate(node_ids):
    ws_pivot.cell(row=1, column=j+2, value=nid)
ws_pivot.cell(row=1, column=len(node_ids)+2, value="Count")

for j in range(1, len(node_ids)+3):
    cell = ws_pivot.cell(row=1, column=j)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Data rows
for i, monster in enumerate(monster_names):
    row = i + 2
    ws_pivot.cell(row=row, column=1, value=monster)
    ws_pivot.cell(row=row, column=1).border = thin_border
    ws_pivot.cell(row=row, column=1).font = Font(bold=True)

    for j, nid in enumerate(node_ids):
        cell = ws_pivot.cell(row=row, column=j+2)
        if nid in pivot[monster]:
            cell.value = pivot[monster][nid]
            cell.number_format = '#,##0'
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right')

    cnt_cell = ws_pivot.cell(row=row, column=len(node_ids)+2, value=occurrences[monster])
    cnt_cell.alignment = Alignment(horizontal='center')
    cnt_cell.border = thin_border

ws_pivot.column_dimensions['A'].width = 28
for j in range(len(node_ids)):
    col_letter = openpyxl.utils.get_column_letter(j+2)
    ws_pivot.column_dimensions[col_letter].width = 14

# ============================================================
# 3. Line Chart with Interpolation
# ============================================================
x_indices = list(range(len(node_ids)))

fig, ax = plt.subplots(figsize=(20, 10))
colors = plt.cm.tab20(np.linspace(0, 1, len(monster_names)))
markers = ['o', 's', '^', 'D', 'v', '<', '>', 'p', '*', 'h', 'H', '8', 'd', 'P']

for idx, monster in enumerate(monster_names):
    actual_x = []
    actual_y = []
    for xi, nid in enumerate(node_ids):
        if nid in pivot[monster]:
            actual_x.append(xi)
            actual_y.append(pivot[monster][nid])

    if len(actual_x) < 1:
        continue

    actual_x = np.array(actual_x)
    actual_y = np.array(actual_y)

    if len(actual_x) >= 2:
        x_full = np.arange(actual_x.min(), actual_x.max() + 1)
        y_interp = np.interp(x_full, actual_x, actual_y)

        color = colors[idx % len(colors)]
        marker = markers[idx % len(markers)]

        ax.plot(x_full, y_interp, '-', color=color, linewidth=1.8, alpha=0.7, label=monster)
        ax.scatter(actual_x, actual_y, c=[color], marker=marker, s=60, zorder=5,
                   edgecolors='black', linewidth=0.5)
    elif len(actual_x) == 1:
        color = colors[idx % len(colors)]
        marker = markers[idx % len(markers)]
        ax.scatter(actual_x, actual_y, c=[color], marker=marker, s=60, zorder=5,
                   edgecolors='black', linewidth=0.5, label=monster)

ax.set_xlabel('Node ID', fontsize=13)
ax.set_ylabel('HP (Million)', fontsize=13)
ax.set_title('Boss HP Trend by Node (Linear Interpolation)', fontsize=16, fontweight='bold')
ax.set_xticks(x_indices)
ax.set_xticklabels(node_ids, rotation=45, ha='right')
ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x/1e6:.0f}M'))
ax.legend(loc='upper left', bbox_to_anchor=(1.01, 1), fontsize=8, ncol=2)
ax.grid(True, alpha=0.3)
ax.set_ylim(bottom=0)

plt.tight_layout()
fig.savefig('../../charts/shiyu/boss_hp_trend_line.png', dpi=150, bbox_inches='tight')
plt.close()
print("\nLine chart saved: boss_hp_trend_line.png")

# ============================================================
# 4. Linear Regression per Monster
# ============================================================
print(f"\n{'Monster':<30s} {'N':>4s} {'Slope(Wan/Node)':>18s} {'Intercept(Wan)':>15s} {'R2':>8s} {'Linearity':>10s} {'Growth':>10s}")
print("-" * 105)

regression_results = []

for monster in monster_names:
    nodes_present = sorted(pivot[monster].keys())
    x_vals = np.array([node_ids.index(n) for n in nodes_present])
    y_vals = np.array([pivot[monster][n] for n in nodes_present])

    n = len(x_vals)
    if n < 2:
        regression_results.append({
            'monster': monster, 'n': n, 'slope': None, 'intercept': None,
            'r2': None, 'nodes': nodes_present, 'x': x_vals, 'y': y_vals,
            'sufficient': False
        })
        print(f"{monster:<30s} {n:>4d}  {'Insufficient data':>35s}")
        continue

    slope, intercept, r_value, p_value, std_err = stats.linregress(x_vals, y_vals)
    r2 = r_value ** 2

    slope_wan = slope / 1e4
    intercept_wan = intercept / 1e4

    if r2 >= 0.8:
        linearity = 'Highly Linear'
    elif r2 >= 0.5:
        linearity = 'Moderate'
    elif r2 >= 0.2:
        linearity = 'Weak'
    else:
        linearity = 'Non-Linear'

    if slope_wan > 1500:
        growth = 'Very Fast'
    elif slope_wan > 800:
        growth = 'Fast'
    elif slope_wan > 300:
        growth = 'Moderate'
    elif slope_wan > 0:
        growth = 'Slow'
    else:
        growth = 'Negative'

    regression_results.append({
        'monster': monster, 'n': n, 'slope': slope, 'intercept': intercept,
        'r2': r2, 'slope_wan': slope_wan, 'intercept_wan': intercept_wan,
        'linearity': linearity, 'growth': growth, 'nodes': nodes_present,
        'x': x_vals, 'y': y_vals, 'sufficient': True
    })

    print(f"{monster:<30s} {n:>4d}  {slope_wan:>18.1f}  {intercept_wan:>15.1f}  {r2:>8.4f}  {linearity:>10s}  {growth:>10s}")

# ============================================================
# 5. Individual Regression Charts
# ============================================================
n_monsters = len([r for r in regression_results if r['sufficient']])
n_cols = 3
n_rows = (n_monsters + n_cols - 1) // n_cols

fig, axes = plt.subplots(n_rows, n_cols, figsize=(6*n_cols, 5*n_rows))
if n_monsters == 1:
    axes = [axes]
else:
    axes = axes.flatten()

plot_idx = 0
for res in regression_results:
    if not res['sufficient']:
        continue

    ax = axes[plot_idx]
    monster = res['monster']
    x_vals = res['x']
    y_vals = res['y']
    slope_wan = res['slope_wan']
    intercept_wan = res['intercept_wan']
    r2 = res['r2']

    ax.scatter(x_vals, y_vals/1e4, c='#2196F3', s=80, zorder=5,
               edgecolors='black', linewidth=0.8, label='Actual Data')

    x_fit = np.linspace(x_vals.min(), x_vals.max(), 100)
    y_fit = res['slope'] * x_fit + res['intercept']
    ax.plot(x_fit, y_fit/1e4, '--', color='#FF5722', linewidth=2,
            label=f'Fit: y={slope_wan:.1f}x+{intercept_wan:.1f}')

    ax.text(0.05, 0.95, f'R2={r2:.4f}\nSlope={slope_wan:.1f}W/node',
            transform=ax.transAxes, fontsize=10, verticalalignment='top',
            fontweight='bold', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))

    ax.set_title(monster, fontsize=11, fontweight='bold')
    ax.set_xlabel('Node Index', fontsize=9)
    ax.set_ylabel('HP (Wan)', fontsize=9)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    node_labels = [node_ids[int(x)] for x in x_vals]
    ax.set_xticks(x_vals)
    ax.set_xticklabels(node_labels, rotation=30, ha='right', fontsize=8)

    plot_idx += 1

for j in range(plot_idx, len(axes)):
    axes[j].set_visible(False)

plt.suptitle('Boss HP Linear Regression Analysis', fontsize=16, fontweight='bold', y=1.01)
plt.tight_layout()
fig.savefig('../../charts/shiyu/boss_hp_regression.png', dpi=150, bbox_inches='tight')
plt.close()
print("Regression charts saved: boss_hp_regression.png")

# ============================================================
# 6. Summary Comparison Chart
# ============================================================
sufficient_results = [r for r in regression_results if r['sufficient']]

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 7))

monsters_short = [r['monster'] for r in sufficient_results]
r2_vals = [r['r2'] for r in sufficient_results]
slopes = [r['slope_wan'] for r in sufficient_results]

# Left: R2 comparison
sorted_idx_r2 = np.argsort(r2_vals)
colors_r2 = ['#4CAF50' if r2 >= 0.8 else '#FF9800' if r2 >= 0.5 else '#f44336'
             for r2 in np.array(r2_vals)[sorted_idx_r2]]

ax1.barh(range(len(sorted_idx_r2)), np.array(r2_vals)[sorted_idx_r2],
         color=colors_r2, edgecolor='black')
ax1.set_yticks(range(len(sorted_idx_r2)))
ax1.set_yticklabels([monsters_short[i] for i in sorted_idx_r2], fontsize=9)
ax1.set_xlabel('R2', fontsize=12)
ax1.set_title('R2 Comparison (Linearity)', fontsize=14, fontweight='bold')
ax1.axvline(x=0.8, color='green', linestyle=':', alpha=0.7, label='Highly Linear (0.8)')
ax1.axvline(x=0.5, color='orange', linestyle=':', alpha=0.7, label='Moderate (0.5)')
ax1.legend(fontsize=9)
ax1.set_xlim(0, 1.05)
ax1.grid(axis='x', alpha=0.3)

# Right: Slope comparison
sorted_idx_slope = np.argsort(slopes)
colors_slope = ['#f44336' if s > 1500 else '#FF9800' if s > 800 else '#2196F3'
                for s in np.array(slopes)[sorted_idx_slope]]

ax2.barh(range(len(sorted_idx_slope)), np.array(slopes)[sorted_idx_slope],
         color=colors_slope, edgecolor='black')
ax2.set_yticks(range(len(sorted_idx_slope)))
ax2.set_yticklabels([monsters_short[i] for i in sorted_idx_slope], fontsize=9)
ax2.set_xlabel('Slope (Wan HP / Node)', fontsize=12)
ax2.set_title('Growth Speed Comparison', fontsize=14, fontweight='bold')
ax2.axvline(x=1500, color='red', linestyle=':', alpha=0.7, label='Very Fast (>1500)')
ax2.axvline(x=800, color='orange', linestyle=':', alpha=0.7, label='Fast (>800)')
ax2.legend(fontsize=9)
ax2.grid(axis='x', alpha=0.3)

plt.tight_layout()
fig.savefig('../../charts/shiyu/boss_hp_summary.png', dpi=150, bbox_inches='tight')
plt.close()
print("Summary chart saved: boss_hp_summary.png")

# ============================================================
# 7. Write Regression Statistics to Excel
# ============================================================
ws_reg = out_wb.create_sheet("Regression Statistics")

reg_headers = ['Monster', 'N Points', 'Nodes', 'Slope(Wan/Node)', 'Intercept(Wan)', 'R2', 'Linearity', 'Growth Speed']
for j, h in enumerate(reg_headers):
    cell = ws_reg.cell(row=1, column=j+1, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

sorted_results = sorted(regression_results,
                        key=lambda r: (r['sufficient'], r.get('r2', 0) if r['sufficient'] else 0),
                        reverse=True)

green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

for i, res in enumerate(sorted_results):
    row = i + 2
    ws_reg.cell(row=row, column=1, value=res['monster']).border = thin_border
    ws_reg.cell(row=row, column=2, value=res['n']).border = thin_border
    ws_reg.cell(row=row, column=2).alignment = Alignment(horizontal='center')

    nodes_str = ', '.join(str(n) for n in res['nodes'])
    ws_reg.cell(row=row, column=3, value=nodes_str).border = thin_border

    if res['sufficient']:
        ws_reg.cell(row=row, column=4, value=round(res['slope_wan'], 1)).border = thin_border
        ws_reg.cell(row=row, column=4).number_format = '#,##0.0'
        ws_reg.cell(row=row, column=4).alignment = Alignment(horizontal='right')

        ws_reg.cell(row=row, column=5, value=round(res['intercept_wan'], 1)).border = thin_border
        ws_reg.cell(row=row, column=5).number_format = '#,##0.0'
        ws_reg.cell(row=row, column=5).alignment = Alignment(horizontal='right')

        ws_reg.cell(row=row, column=6, value=round(res['r2'], 4)).border = thin_border
        ws_reg.cell(row=row, column=6).number_format = '0.0000'
        ws_reg.cell(row=row, column=6).alignment = Alignment(horizontal='center')

        ws_reg.cell(row=row, column=7, value=res['linearity']).border = thin_border
        ws_reg.cell(row=row, column=8, value=res['growth']).border = thin_border

        if res['linearity'] == 'Highly Linear':
            ws_reg.cell(row=row, column=7).fill = green_fill
        elif res['linearity'] == 'Moderate':
            ws_reg.cell(row=row, column=7).fill = yellow_fill
        else:
            ws_reg.cell(row=row, column=7).fill = red_fill

        if res['growth'] in ('Very Fast', 'Fast'):
            ws_reg.cell(row=row, column=8).fill = red_fill
        elif res['growth'] == 'Moderate':
            ws_reg.cell(row=row, column=8).fill = yellow_fill
        else:
            ws_reg.cell(row=row, column=8).fill = green_fill
    else:
        ws_reg.cell(row=row, column=4, value='N/A').border = thin_border
        ws_reg.cell(row=row, column=4).fill = red_fill
        for j in range(4, 9):
            ws_reg.cell(row=row, column=j).border = thin_border

    for j in range(1, 9):
        ws_reg.cell(row=row, column=j).alignment = Alignment(
            horizontal='center' if j != 1 else 'left')

col_widths = [28, 10, 40, 18, 16, 10, 14, 16]
for j, w in enumerate(col_widths):
    ws_reg.column_dimensions[openpyxl.utils.get_column_letter(j+1)].width = w

# ============================================================
# 8. Analysis Summary Sheet
# ============================================================
ws_summary = out_wb.create_sheet("Summary")

summary_lines = [
    "Boss HP Inflation Analysis Summary",
    "",
    "=== 1. Data Overview ===",
    f"Node range: {node_ids[0]} ~ {node_ids[-1]} ({len(node_ids)} nodes)",
    f"Boss types: {len(monster_names)}",
    f"Total boss entries: {len(bosses)}",
    f"Avg occurrences per boss: {sum(occurrences.values())/len(occurrences):.1f}",
    "",
    "=== 2. Growth Speed Ranking (Slope = HP increase per node, in Wan) ===",
]

speed_rank = sorted(sufficient_results, key=lambda r: -r['slope_wan'])
for rank, res in enumerate(speed_rank, 1):
    summary_lines.append(f"  {rank}. {res['monster']}: {res['slope_wan']:.1f} Wan/node (R2={res['r2']:.4f})")

summary_lines.append("")
summary_lines.append("=== 3. Linearity Ranking (higher R2 = more linear) ===")

r2_rank = sorted(sufficient_results, key=lambda r: -r['r2'])
for rank, res in enumerate(r2_rank, 1):
    summary_lines.append(f"  {rank}. {res['monster']}: R2={res['r2']:.4f} (Slope={res['slope_wan']:.1f})")

summary_lines.append("")
summary_lines.append("=== 4. Comprehensive Analysis ===")

fastest = speed_rank[0]
most_linear = r2_rank[0]
summary_lines.append(f"Fastest growth: {fastest['monster']} (Slope={fastest['slope_wan']:.1f}W/node, R2={fastest['r2']:.4f})")
summary_lines.append(f"Most linear: {most_linear['monster']} (R2={most_linear['r2']:.4f}, Slope={most_linear['slope_wan']:.1f}W/node)")

hl_fast = [r for r in sufficient_results if r['linearity'] == 'Highly Linear' and r['growth'] in ('Very Fast', 'Fast')]
if hl_fast:
    summary_lines.append(f"Highly linear & fast growth: {', '.join(r['monster'] for r in hl_fast)}")
    summary_lines.append("  -> These bosses have predictable, sustained high inflation - need attention!")

nonlinear = [r for r in sufficient_results if r['linearity'] in ('Weak', 'Non-Linear')]
if nonlinear:
    summary_lines.append(f"Erratic growth: {', '.join(r['monster'] for r in nonlinear)}")
    summary_lines.append("  -> HP changes are unstable, possibly affected by special design factors")

neg_growth = [r for r in sufficient_results if r['slope'] < 0]
if neg_growth:
    summary_lines.append(f"Negative growth: {', '.join(r['monster'] for r in neg_growth)}")
    summary_lines.append("  -> HP decreases over versions, may be nerfed or rotated to lower difficulty")

for i, line in enumerate(summary_lines):
    cell = ws_summary.cell(row=i+1, column=1, value=line)
    if i == 0:
        cell.font = Font(bold=True, size=14)
    elif line.startswith("==="):
        cell.font = Font(bold=True, size=12)

ws_summary.column_dimensions['A'].width = 90

# Save
output_path = '../../charts/shiyu/boss_hp_analysis.xlsx'
out_wb.save(output_path)
print(f"\nAnalysis Excel saved: {output_path}")

# --- Chart 5: R2 vs Growth Rate Scatter (same style as 危局 chart_r2_vs_growth) ---
fig, ax = plt.subplots(figsize=(14, 8))

# Compute growth rates for scatter
scatter_r2 = np.array([r['r2'] for r in sufficient_results])
scatter_slopes = np.array([r['slope_wan'] for r in sufficient_results])
scatter_n = np.array([r['n'] for r in sufficient_results])
# Total growth as color: for each boss, (last_hp - first_hp) / first_hp * 100
scatter_growth = []
for r in sufficient_results:
    y_vals = r['y']
    total_g = (y_vals[-1] - y_vals[0]) / y_vals[0] * 100
    scatter_growth.append(total_g)

sc = ax.scatter(scatter_r2, scatter_slopes,
                s=scatter_n * 20,
                c=scatter_growth,
                cmap='YlOrRd', alpha=0.8, edgecolors='black', linewidth=0.5)

# Annotate each point
for i, r in enumerate(sufficient_results):
    ax.annotate(r['monster'],
               xy=(scatter_r2[i], scatter_slopes[i]),
               xytext=(5, 5), textcoords='offset points',
               fontsize=7, alpha=0.8)

# Median reference lines
median_r2 = np.median(scatter_r2)
median_slope = np.median(scatter_slopes)
ax.axhline(y=median_slope, color='blue', linestyle='--', alpha=0.5,
          label=f'Growth Median: {median_slope:.1f} Wan/node')
ax.axvline(x=median_r2, color='green', linestyle='--', alpha=0.5,
          label=f'R2 Median: {median_r2:.3f}')

# Quadrant labels
ax.text(0.98, 0.98, 'High Linear + Fast Growth\n(Most Notable)', transform=ax.transAxes,
       fontsize=8, ha='right', va='top', alpha=0.6)
ax.text(0.02, 0.98, 'Low Linear + Fast Growth\n(Volatile but Fast)', transform=ax.transAxes,
       fontsize=8, ha='left', va='top', alpha=0.6)
ax.text(0.98, 0.02, 'High Linear + Slow Growth\n(Stable & Slow)', transform=ax.transAxes,
       fontsize=8, ha='right', va='bottom', alpha=0.6)
ax.text(0.02, 0.02, 'Low Linear + Slow Growth\n(Volatile & Slow)', transform=ax.transAxes,
       fontsize=8, ha='left', va='bottom', alpha=0.6)

ax.set_xlabel('R2 (Linear Fit)', fontsize=12)
ax.set_ylabel('Growth Rate (Wan HP / Node)', fontsize=12)
ax.set_title('Shiyu Boss HP Inflation: Linearity vs Growth Rate\n'
             '(Point Size = Occurrence Count, Color = Total Growth %)',
            fontsize=14, fontweight='bold')
cbar = plt.colorbar(sc, ax=ax)
cbar.set_label('Total Growth (%)', fontsize=10)
ax.legend(fontsize=9)
ax.grid(True, alpha=0.3)

plt.tight_layout()
fig.savefig('../../charts/shiyu/chart_r2_vs_growth.png', dpi=150, bbox_inches='tight')
plt.close()
print("Chart 5/5 saved: chart_r2_vs_growth.png (R2 vs Growth Rate)")

print("\n" + "=" * 60)
print("DONE! Output files:")
print("  1. boss_hp_analysis.xlsx (Pivot Table + Regression Stats + Summary)")
print("  2. boss_hp_trend_line.png")
print("  3. boss_hp_regression.png")
print("  4. boss_hp_summary.png")
