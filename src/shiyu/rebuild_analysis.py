import openpyxl
import numpy as np
from scipy import stats
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from collections import defaultdict
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import warnings
warnings.filterwarnings('ignore')

plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# ============================================================
# 1. Load data
# ============================================================
wb = openpyxl.load_workbook('式舆怪物血量数据.xlsx')
ws = wb[wb.sheetnames[0]]

bosses = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    node_id = row[0]
    monster = row[3]
    hp_str = row[4]
    if hp_str is None:
        continue
    hp = int(str(hp_str).replace(',', ''))
    if hp > 12_600_000:
        bosses.append((node_id, monster, hp))

node_ids = sorted(set(b[0] for b in bosses))
monster_names = sorted(set(b[1] for b in bosses))
print(f"Nodes: {len(node_ids)}, Boss types: {len(monster_names)}")
print(f"Node IDs: {node_ids}")

pivot = defaultdict(dict)
for nid, monster, hp in bosses:
    pivot[monster][nid] = hp

occurrences = {m: len(pivot[m]) for m in monster_names}

for m in monster_names:
    print(f"  {m}: {occurrences[m]} occurrences -> {sorted(pivot[m].keys())}")

# ============================================================
# 2. Styles
# ============================================================
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
bold_font = Font(bold=True)
normal_font = Font(size=11)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center')
right_align = Alignment(horizontal='right')
left_align = Alignment(horizontal='left')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

def sw(ws, row, col, value, font=None, fill=None, alignment=None, number_format=None, border=None):
    """Safe write - no formulas, pure values only."""
    cell = ws.cell(row=row, column=col)
    # Ensure value is NOT a formula string (starting with =)
    if isinstance(value, str) and value.lstrip().startswith('='):
        cell.value = "'" + value  # prefix with single quote to force text
    else:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    return cell

# ============================================================
# 3. Pivot Table (Sheet 1)
# ============================================================
out_wb = openpyxl.Workbook()
ws_pivot = out_wb.active
ws_pivot.title = "Pivot Table"

sw(ws_pivot, 1, 1, "Monster", font=header_font, fill=header_fill, alignment=center_align, border=thin_border)
for j, nid in enumerate(node_ids):
    sw(ws_pivot, 1, j+2, nid, font=header_font, fill=header_fill, alignment=center_align, border=thin_border)
sw(ws_pivot, 1, len(node_ids)+2, "Count", font=header_font, fill=header_fill, alignment=center_align, border=thin_border)

for i, monster in enumerate(monster_names):
    row = i + 2
    sw(ws_pivot, row, 1, monster, font=bold_font, alignment=left_align, border=thin_border)
    for j, nid in enumerate(node_ids):
        if nid in pivot[monster]:
            sw(ws_pivot, row, j+2, pivot[monster][nid], number_format='#,##0', alignment=right_align, border=thin_border)
        else:
            sw(ws_pivot, row, j+2, None, border=thin_border)
    sw(ws_pivot, row, len(node_ids)+2, occurrences[monster], alignment=center_align, border=thin_border)

ws_pivot.column_dimensions['A'].width = 28
for j in range(len(node_ids)):
    ws_pivot.column_dimensions[openpyxl.utils.get_column_letter(j+2)].width = 14
ws_pivot.freeze_panes = 'A2'

# ============================================================
# 4. Linear Regression
# ============================================================
regression_results = []

for monster in monster_names:
    nodes_present = sorted(pivot[monster].keys())
    x_vals = np.array([node_ids.index(n) for n in nodes_present])
    y_vals = np.array([pivot[monster][n] for n in nodes_present])
    n = len(x_vals)

    if n < 3:
        regression_results.append({
            'monster': monster, 'n': n, 'slope': None, 'intercept': None,
            'r2': None, 'nodes': nodes_present, 'x': x_vals, 'y': y_vals,
            'sufficient': False
        })
        continue

    slope, intercept, r_value, p_value, std_err = stats.linregress(x_vals, y_vals)
    r2 = r_value ** 2
    slope_wan = slope / 1e4
    intercept_wan = intercept / 1e4

    if r2 >= 0.8:
        linearity = 'Highly Linear'
    elif r2 >= 0.5:
        linearity = 'Moderate'
    elif r2 >= 0.2:
        linearity = 'Weak'
    else:
        linearity = 'Non-Linear'

    if slope_wan > 300:
        growth = 'Very Fast'
    elif slope_wan > 150:
        growth = 'Fast'
    elif slope_wan > 80:
        growth = 'Moderate'
    else:
        growth = 'Slow'

    regression_results.append({
        'monster': monster, 'n': n, 'slope': slope, 'intercept': intercept,
        'r2': r2, 'slope_wan': slope_wan, 'intercept_wan': intercept_wan,
        'linearity': linearity, 'growth': growth, 'nodes': nodes_present,
        'x': x_vals, 'y': y_vals, 'sufficient': True
    })

# ============================================================
# 5. Regression Statistics Sheet (Sheet 2)
# ============================================================
ws_reg = out_wb.create_sheet("Regression Stats")

reg_headers = ['Monster', 'Data Points', 'Nodes', 'Slope (Wan/Node)', 'Intercept (Wan)', 'R2', 'Linearity', 'Growth Speed']
for j, h in enumerate(reg_headers):
    sw(ws_reg, 1, j+1, h, font=header_font, fill=header_fill, alignment=center_align, border=thin_border)

sorted_results = sorted(regression_results,
                        key=lambda r: (r['sufficient'], r.get('r2', 0) if r['sufficient'] else 0),
                        reverse=True)

for i, res in enumerate(sorted_results):
    row = i + 2
    sw(ws_reg, row, 1, res['monster'], font=bold_font, alignment=left_align, border=thin_border)
    sw(ws_reg, row, 2, res['n'], alignment=center_align, border=thin_border)

    nodes_str = ', '.join(str(n) for n in res['nodes'])
    sw(ws_reg, row, 3, nodes_str, alignment=left_align, border=thin_border)

    if res['sufficient']:
        sw(ws_reg, row, 4, round(res['slope_wan'], 1), number_format='#,##0.0', alignment=right_align, border=thin_border)
        sw(ws_reg, row, 5, round(res['intercept_wan'], 1), number_format='#,##0.0', alignment=right_align, border=thin_border)
        sw(ws_reg, row, 6, round(res['r2'], 4), number_format='0.0000', alignment=center_align, border=thin_border)

        lin_fill = green_fill if res['linearity'] == 'Highly Linear' else (yellow_fill if res['linearity'] == 'Moderate' else red_fill)
        sw(ws_reg, row, 7, res['linearity'], fill=lin_fill, alignment=center_align, border=thin_border)

        grow_fill = red_fill if res['growth'] in ('Very Fast', 'Fast') else (yellow_fill if res['growth'] == 'Moderate' else green_fill)
        sw(ws_reg, row, 8, res['growth'], fill=grow_fill, alignment=center_align, border=thin_border)
    else:
        for j in range(4, 9):
            sw(ws_reg, row, j, 'N/A', fill=red_fill, alignment=center_align, border=thin_border)

col_widths = [28, 12, 40, 18, 16, 10, 14, 16]
for j, w in enumerate(col_widths):
    ws_reg.column_dimensions[openpyxl.utils.get_column_letter(j+1)].width = w
ws_reg.freeze_panes = 'A2'

# ============================================================
# 6. Summary Sheet (Sheet 3) - Plain text, no formulas
# ============================================================
ws_summary = out_wb.create_sheet("Summary")

sufficient_results = [r for r in regression_results if r['sufficient']]
speed_rank = sorted(sufficient_results, key=lambda r: -r['slope_wan'])
r2_rank = sorted(sufficient_results, key=lambda r: -r['r2'])

row = 1
sw(ws_summary, row, 1, "Boss HP Inflation Analysis Summary", font=title_font); row += 2

sw(ws_summary, row, 1, "=== 1. Data Overview ===", font=section_font); row += 1
lines_overview = [
    "Node range: %d - %d (%d nodes)" % (node_ids[0], node_ids[-1], len(node_ids)),
    "Boss types: %d" % len(monster_names),
    "Total boss entries: %d" % len(bosses),
    "Avg occurrences per boss: %.1f" % (sum(occurrences.values())/len(occurrences)),
    "Bosses with enough data (>=3 points) for regression: %d" % len(sufficient_results),
    "Bosses with insufficient data (<3 points): %d" % (len(regression_results) - len(sufficient_results)),
]
for line in lines_overview:
    sw(ws_summary, row, 1, line, font=normal_font); row += 1
row += 1

sw(ws_summary, row, 1, "=== 2. Growth Speed Ranking ===", font=section_font); row += 1
sw(ws_summary, row, 1, "Slope = HP increase per node step, in Wan (10,000 HP)", font=normal_font); row += 1
for rank, res in enumerate(speed_rank, 1):
    line = "  %d. %s: %.1f Wan/node, R2=%.4f, %s" % (rank, res['monster'], res['slope_wan'], res['r2'], res['growth'])
    sw(ws_summary, row, 1, line, font=normal_font); row += 1
row += 1

sw(ws_summary, row, 1, "=== 3. Linearity Ranking ===", font=section_font); row += 1
sw(ws_summary, row, 1, "Higher R2 = more linear (predictable) growth pattern", font=normal_font); row += 1
for rank, res in enumerate(r2_rank, 1):
    line = "  %d. %s: R2=%.4f, Slope=%.1f Wan/node, %s" % (rank, res['monster'], res['r2'], res['slope_wan'], res['linearity'])
    sw(ws_summary, row, 1, line, font=normal_font); row += 1
row += 1

sw(ws_summary, row, 1, "=== 4. Comprehensive Analysis ===", font=section_font); row += 1

fastest = speed_rank[0]
most_linear = r2_rank[0]
sw(ws_summary, row, 1, "Fastest growth: %s (Slope=%.1f Wan/node, R2=%.4f)" % (fastest['monster'], fastest['slope_wan'], fastest['r2']), font=bold_font); row += 1
sw(ws_summary, row, 1, "Most linear: %s (R2=%.4f, Slope=%.1f Wan/node)" % (most_linear['monster'], most_linear['r2'], most_linear['slope_wan']), font=bold_font); row += 2

hl_fast = [r for r in sufficient_results if r['linearity'] == 'Highly Linear' and r['growth'] in ('Very Fast', 'Fast')]
if hl_fast:
    names = ', '.join(r['monster'] for r in hl_fast)
    sw(ws_summary, row, 1, "Highly linear AND fast growth: %s" % names, font=bold_font); row += 1
    sw(ws_summary, row, 1, "  -> These bosses show steady, predictable inflation at a high rate. Priority monitoring targets.", font=normal_font); row += 2

nonlinear = [r for r in sufficient_results if r['linearity'] in ('Weak', 'Non-Linear')]
if nonlinear:
    names = ', '.join(r['monster'] for r in nonlinear)
    sw(ws_summary, row, 1, "Erratic growth (non-linear): %s" % names, font=normal_font); row += 1
    sw(ws_summary, row, 1, "  -> These bosses have irregular HP changes, possibly due to redesign or special mechanics.", font=normal_font); row += 2

moderate_lin = [r for r in sufficient_results if r['linearity'] == 'Moderate']
if moderate_lin:
    names = ', '.join(r['monster'] for r in moderate_lin)
    sw(ws_summary, row, 1, "Moderately linear: %s" % names, font=normal_font); row += 1
    sw(ws_summary, row, 1, "  -> These show some trend but with notable fluctuations between versions.", font=normal_font); row += 2

# Note about insufficient data
two_pt = [r for r in regression_results if not r['sufficient'] and r['n'] == 2]
one_pt = [r for r in regression_results if r['n'] < 2]

sw(ws_summary, row, 1, "=== 5. Data Limitations ===", font=section_font); row += 1

if two_pt:
    sw(ws_summary, row, 1, "Bosses with only 2 data points (regression trivially perfect R2=1.0, not meaningful):", font=normal_font); row += 1
    for r in two_pt:
        nodes_str = ', '.join(str(n) for n in r['nodes'])
        line = "  - %s (nodes: %s)" % (r['monster'], nodes_str)
        sw(ws_summary, row, 1, line, font=normal_font); row += 1
    row += 1

if one_pt:
    sw(ws_summary, row, 1, "Bosses with only 1 data point (no trend possible):", font=normal_font); row += 1
    for r in one_pt:
        line = "  - %s (node: %d)" % (r['monster'], r['nodes'][0])
        sw(ws_summary, row, 1, line, font=normal_font); row += 1

ws_summary.column_dimensions['A'].width = 100

# ============================================================
# 7. Charts
# ============================================================
x_indices = list(range(len(node_ids)))
colors = plt.cm.tab20(np.linspace(0, 1, len(monster_names)))
markers = ['o', 's', '^', 'D', 'v', '<', '>', 'p', '*', 'h', 'H', '8', 'd', 'P']

# --- Chart 1: Line Chart with Interpolation ---
fig, ax = plt.subplots(figsize=(20, 10))

for idx, monster in enumerate(monster_names):
    actual_x, actual_y = [], []
    for xi, nid in enumerate(node_ids):
        if nid in pivot[monster]:
            actual_x.append(xi)
            actual_y.append(pivot[monster][nid])
    if not actual_x:
        continue
    actual_x, actual_y = np.array(actual_x), np.array(actual_y)
    color, marker = colors[idx % len(colors)], markers[idx % len(markers)]

    if len(actual_x) >= 2:
        x_full = np.arange(actual_x.min(), actual_x.max() + 1)
        y_interp = np.interp(x_full, actual_x, actual_y)
        ax.plot(x_full, y_interp, '-', color=color, linewidth=1.8, alpha=0.7, label=monster)
        ax.scatter(actual_x, actual_y, c=[color], marker=marker, s=60, zorder=5, edgecolors='black', linewidth=0.5)
    else:
        ax.scatter(actual_x, actual_y, c=[color], marker=marker, s=60, zorder=5, edgecolors='black', linewidth=0.5, label=monster)

ax.set_xlabel('Node ID', fontsize=13)
ax.set_ylabel('HP (Million)', fontsize=13)
ax.set_title('Boss HP Trend by Node (Linear Interpolation)', fontsize=16, fontweight='bold')
ax.set_xticks(x_indices)
ax.set_xticklabels(node_ids, rotation=45, ha='right')
ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x/1e6:.0f}M'))
ax.legend(loc='upper left', bbox_to_anchor=(1.01, 1), fontsize=8, ncol=2)
ax.grid(True, alpha=0.3)
ax.set_ylim(bottom=0)
plt.tight_layout()
fig.savefig('../../charts/shiyu/boss_hp_trend_line.png', dpi=150, bbox_inches='tight')
plt.close()
print("Chart 1/4 saved: boss_hp_trend_line.png")

# --- Chart 2: Individual Regression Charts ---
n_reg = len(sufficient_results)
n_cols = 3
n_rows = max(1, (n_reg + n_cols - 1) // n_cols)

fig, axes = plt.subplots(n_rows, n_cols, figsize=(6*n_cols, 5*n_rows))
if n_reg == 1:
    axes = [axes]
else:
    axes = axes.flatten()

for plot_idx, res in enumerate(sufficient_results):
    ax = axes[plot_idx]
    x_vals, y_vals = res['x'], res['y']

    ax.scatter(x_vals, y_vals/1e4, c='#2196F3', s=80, zorder=5, edgecolors='black', linewidth=0.8, label='Actual')
    x_fit = np.linspace(x_vals.min(), x_vals.max(), 100)
    y_fit = res['slope'] * x_fit + res['intercept']
    ax.plot(x_fit, y_fit/1e4, '--', color='#FF5722', linewidth=2,
            label='y=%.1fx+%.1f' % (res['slope_wan'], res['intercept_wan']))

    ax.text(0.05, 0.95, 'R2=%.4f\nSlope=%.1fW/node' % (res['r2'], res['slope_wan']),
            transform=ax.transAxes, fontsize=10, verticalalignment='top', fontweight='bold',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
    ax.set_title(res['monster'], fontsize=11, fontweight='bold')
    ax.set_xlabel('Node Index', fontsize=9)
    ax.set_ylabel('HP (Wan)', fontsize=9)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    node_labels = [node_ids[int(x)] for x in x_vals]
    ax.set_xticks(x_vals)
    ax.set_xticklabels(node_labels, rotation=30, ha='right', fontsize=8)

for j in range(n_reg, len(axes)):
    axes[j].set_visible(False)

plt.suptitle('Boss HP Linear Regression (>=3 data points)', fontsize=16, fontweight='bold', y=1.01)
plt.tight_layout()
fig.savefig('../../charts/shiyu/boss_hp_regression.png', dpi=150, bbox_inches='tight')
plt.close()
print("Chart 2/4 saved: boss_hp_regression.png")

# --- Chart 3: Summary Comparison ---
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 7))

monsters_short = [r['monster'] for r in sufficient_results]
r2_vals = [r['r2'] for r in sufficient_results]
slopes = [r['slope_wan'] for r in sufficient_results]

sort_r2 = np.argsort(r2_vals)
colors_r2 = ['#4CAF50' if v >= 0.8 else '#FF9800' if v >= 0.5 else '#f44336' for v in np.array(r2_vals)[sort_r2]]
ax1.barh(range(len(sort_r2)), np.array(r2_vals)[sort_r2], color=colors_r2, edgecolor='black')
ax1.set_yticks(range(len(sort_r2)))
ax1.set_yticklabels([monsters_short[i] for i in sort_r2], fontsize=9)
ax1.set_xlabel('R2', fontsize=12)
ax1.set_title('R2 Comparison (Linearity)', fontsize=14, fontweight='bold')
ax1.axvline(x=0.8, color='green', linestyle=':', alpha=0.7, label='Highly Linear (0.8)')
ax1.axvline(x=0.5, color='orange', linestyle=':', alpha=0.7, label='Moderate (0.5)')
ax1.legend(fontsize=9)
ax1.set_xlim(0, 1.05)
ax1.grid(axis='x', alpha=0.3)

sort_slope = np.argsort(slopes)
colors_slope = ['#f44336' if s > 300 else '#FF9800' if s > 150 else '#2196F3' for s in np.array(slopes)[sort_slope]]
ax2.barh(range(len(sort_slope)), np.array(slopes)[sort_slope], color=colors_slope, edgecolor='black')
ax2.set_yticks(range(len(sort_slope)))
ax2.set_yticklabels([monsters_short[i] for i in sort_slope], fontsize=9)
ax2.set_xlabel('Slope (Wan HP / Node)', fontsize=12)
ax2.set_title('Growth Speed Comparison', fontsize=14, fontweight='bold')
ax2.axvline(x=300, color='red', linestyle=':', alpha=0.7, label='Very Fast (>300)')
ax2.axvline(x=150, color='orange', linestyle=':', alpha=0.7, label='Fast (>150)')
ax2.legend(fontsize=9)
ax2.grid(axis='x', alpha=0.3)

plt.tight_layout()
fig.savefig('../../charts/shiyu/boss_hp_summary.png', dpi=150, bbox_inches='tight')
plt.close()
print("Chart 3/4 saved: boss_hp_summary.png")

# --- Chart 4: Combined Overview ---
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(20, 16))

for idx, monster in enumerate(monster_names):
    actual_x, actual_y = [], []
    for xi, nid in enumerate(node_ids):
        if nid in pivot[monster]:
            actual_x.append(xi)
            actual_y.append(pivot[monster][nid])
    if not actual_x:
        continue
    actual_x, actual_y = np.array(actual_x), np.array(actual_y)
    color, marker = colors[idx % len(colors)], markers[idx % len(markers)]
    if len(actual_x) >= 2:
        x_full = np.arange(actual_x.min(), actual_x.max() + 1)
        y_interp = np.interp(x_full, actual_x, actual_y)
        ax1.plot(x_full, y_interp, '-', color=color, linewidth=1.8, alpha=0.7, label=monster)
        ax1.scatter(actual_x, actual_y, c=[color], marker=marker, s=60, zorder=5, edgecolors='black', linewidth=0.5)
    else:
        ax1.scatter(actual_x, actual_y, c=[color], marker=marker, s=60, zorder=5, edgecolors='black', linewidth=0.5, label=monster)

ax1.set_xlabel('Node ID', fontsize=13)
ax1.set_ylabel('HP (Million)', fontsize=13)
ax1.set_title('Boss HP Trend (Linear Interpolation between Actual Points)', fontsize=14, fontweight='bold')
ax1.set_xticks(x_indices)
ax1.set_xticklabels(node_ids, rotation=45, ha='right')
ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x/1e6:.0f}M'))
ax1.legend(loc='upper left', bbox_to_anchor=(1.01, 1), fontsize=8, ncol=2)
ax1.grid(True, alpha=0.3)
ax1.set_ylim(bottom=0)

for idx, res in enumerate(sufficient_results):
    x_vals, y_vals = res['x'], res['y']
    x_full = np.arange(x_vals.min(), x_vals.max() + 1)
    y_fit = res['slope'] * x_full + res['intercept']
    color, marker = colors[idx % len(colors)], markers[idx % len(markers)]
    ax2.plot(x_full, y_fit/1e4, '-', color=color, linewidth=2, alpha=0.8,
             label="%s (R2=%.3f)" % (res['monster'], res['r2']))
    ax2.scatter(x_vals, y_vals/1e4, c=[color], marker=marker, s=60, zorder=5, edgecolors='black', linewidth=0.5)

ax2.set_xlabel('Node ID', fontsize=13)
ax2.set_ylabel('HP (Wan)', fontsize=13)
ax2.set_title('Boss HP Linear Regression Lines (>= 3 data points)', fontsize=14, fontweight='bold')
ax2.set_xticks(x_indices)
ax2.set_xticklabels(node_ids, rotation=45, ha='right')
ax2.legend(loc='upper left', bbox_to_anchor=(1.01, 1), fontsize=9, ncol=2)
ax2.grid(True, alpha=0.3)

plt.tight_layout()
fig.savefig('../../charts/shiyu/boss_hp_combined.png', dpi=150, bbox_inches='tight')
plt.close()
print("Chart 4/4 saved: boss_hp_combined.png")

# ============================================================
# 8. Save Excel
# ============================================================
output_path = '../../charts/shiyu/boss_hp_analysis.xlsx'
# Remove old file if exists
import os
if os.path.exists(output_path):
    os.remove(output_path)
out_wb.save(output_path)
print("\nExcel saved: %s" % output_path)
print("DONE - All files regenerated successfully!")
